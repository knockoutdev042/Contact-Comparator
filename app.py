"""
app.py
-------------------------
Streamlit web app for the Excluded Contacts Extractor.

Ports excluded_contacts_gui.py's core logic to the web: flexible
phone/email column detection, email-priority matching against a
Salesforce file (email checked first when present, phone as
fallback), and a styled multi-sheet Excel export. The Tkinter GUI
is replaced by Streamlit widgets; matching/normalization logic is
unchanged from the original.

Run locally:
    streamlit run app.py
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auth import render_logout_button, require_login

# ── Normalization Helpers ─────────────────────────────────


def normalize_phone(p):
    if pd.isna(p):
        return None

    p = str(p)
    p = re.sub(r"\s+", "", p)
    p = p.replace("'", "")

    try:
        if "E+" in p.upper():
            p = str(int(float(p)))
    except Exception:
        pass

    digits = re.sub(r"\D", "", p)
    return f"+{digits}" if digits else None


def normalize_email(email):
    if pd.isna(email):
        return None

    email = str(email)
    email = re.sub(r"\s+", "", email)
    email = email.lower().strip()

    return email if "@" in email else None


def find_column(df, possible_names):
    for col in df.columns:
        clean_col = col.strip().lower()
        for name in possible_names:
            if clean_col == name.strip().lower():
                return col
    return None


PHONE_ALIASES = ["Phone", "Mobile", "Phone Number", "Contact Number"]
EMAIL_ALIASES = ["Email", "Email Address", "E-mail"]


def read_any(uploaded_file):
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file, encoding="utf-8-sig")
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding="latin1")
    else:
        return pd.read_excel(uploaded_file)


# ── Excel Styling ─────────────────────────────────


def style_workbook(buffer):
    buffer.seek(0)
    wb = load_workbook(buffer)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Calibri", size=9)

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ── Core Logic ────────────────────────────────────


def run_extraction(contact_files, sf_file, log):
    log("Loading Contact Us files...")

    contact_dfs = []

    for f in contact_files:
        log(f"  → {f.name}")

        try:
            df = read_any(f)
        except Exception as e:
            log(f"❌ Failed to read file: {e}")
            continue

        if df.empty:
            log(f"⚠ {f.name} is empty")
            continue

        df.columns = df.columns.str.strip()

        phone_col = find_column(df, PHONE_ALIASES)
        email_col = find_column(df, EMAIL_ALIASES)

        if not phone_col and not email_col:
            log(f"❌ Skipping {f.name}: No Phone or Email column found")
            continue

        df["_phone_norm"] = df[phone_col].apply(normalize_phone) if phone_col else None
        df["_email_norm"] = df[email_col].apply(normalize_email) if email_col else None

        contact_dfs.append(df)

    if not contact_dfs:
        raise Exception("No valid Contact files found.")

    contact = pd.concat(contact_dfs, ignore_index=True)
    log(f"✔ Total Contact rows merged: {len(contact):,}")

    # ── Load Salesforce File ─────────────────────────────

    log("Loading Salesforce file...")

    try:
        sf = read_any(sf_file)
    except Exception as e:
        raise Exception(f"Failed to read Salesforce file: {e}")

    sf.columns = sf.columns.str.strip()

    sf_phone_col = find_column(sf, PHONE_ALIASES)
    sf_email_col = find_column(sf, EMAIL_ALIASES)

    if not sf_phone_col and not sf_email_col:
        raise Exception("No Phone or Email column found in Salesforce file")

    log(f"✔ Salesforce rows loaded: {len(sf):,}")

    sf["_phone_norm"] = sf[sf_phone_col].apply(normalize_phone) if sf_phone_col else None
    sf["_email_norm"] = sf[sf_email_col].apply(normalize_email) if sf_email_col else None

    # Remove invalid rows
    contact = contact[contact["_phone_norm"].notna() | contact["_email_norm"].notna()]
    sf = sf[sf["_phone_norm"].notna() | sf["_email_norm"].notna()]

    # Remove duplicates
    contact = contact.drop_duplicates(subset=["_phone_norm", "_email_norm"])
    sf = sf.drop_duplicates(subset=["_phone_norm", "_email_norm"])

    sf_phones = set(sf["_phone_norm"].dropna())
    sf_emails = set(sf["_email_norm"].dropna())

    log(f"✔ Unique SF phones : {len(sf_phones):,}")
    log(f"✔ Unique SF emails : {len(sf_emails):,}")

    # ── Matching Logic ─────────────────────────────────

    log("Comparing Contact data against Salesforce...")

    email_exists = contact["_email_norm"].notna()
    email_match = contact["_email_norm"].isin(sf_emails)
    phone_match = contact["_phone_norm"].isin(sf_phones)

    # If email exists -> ONLY email comparison. Else -> fallback to phone.
    matched_mask = (email_exists & email_match) | (~email_exists & phone_match)

    matched = contact[matched_mask].copy()
    excluded = contact[~matched_mask].copy()

    matched_count = int(matched_mask.sum())

    helper_cols = ["_phone_norm", "_email_norm"]
    matched.drop(columns=helper_cols, inplace=True, errors="ignore")
    excluded.drop(columns=helper_cols, inplace=True, errors="ignore")

    log("\nResults:")
    log(f"  Total Contact rows : {len(contact):,}")
    log(f"  Matched rows       : {matched_count:,}")
    log(f"  Excluded rows      : {len(excluded):,}")

    return {
        "contact_total": len(contact),
        "matched": matched,
        "excluded": excluded,
        "matched_count": matched_count,
        "excluded_count": len(excluded),
    }


# ── Styling ────────────────────────────────────

STYLE = """
<style>
.ece-banner {
    background: #16213e;
    border-radius: 14px;
    padding: 22px 28px;
    margin-bottom: 22px;
    display: flex;
    align-items: center;
    gap: 16px;
}
.ece-banner .ece-icon { font-size: 2.2rem; line-height: 1; }
.ece-banner .ece-title { color: #e94560; font-size: 1.5rem; font-weight: 700; letter-spacing: .01em; }
.ece-banner .ece-subtitle { color: #a0a0c0; font-size: 0.88rem; margin-top: 3px; }

.stat-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 14px;
    margin: 14px 0 22px;
}
@media (max-width: 900px) {
    .stat-grid { grid-template-columns: repeat(2, 1fr); }
}
.stat-card {
    border-radius: 12px;
    padding: 16px 18px;
    border: 1px solid rgba(11,11,11,0.10);
    border-left: 4px solid #2a78d6;
    background: #fcfcfb;
}
.stat-card .stat-icon { font-size: 1.3rem; }
.stat-card .stat-label {
    font-size: 0.78rem;
    color: #52514e;
    margin-top: 6px;
    font-weight: 600;
    letter-spacing: .03em;
    text-transform: uppercase;
}
.stat-card .stat-value { font-size: 1.9rem; font-weight: 700; color: #0b0b0b; margin-top: 2px; }
.stat-card.role-good { border-left-color: #0ca30c; }
.stat-card.role-serious { border-left-color: #ec835a; }
.stat-card.role-accent { border-left-color: #e94560; }

@media (prefers-color-scheme: dark) {
    .stat-card { background: #1a1a19; border-color: rgba(255,255,255,0.10); }
    .stat-card .stat-label { color: #c3c2b7; }
    .stat-card .stat-value { color: #ffffff; }
    .stat-card.role-neutral { border-left-color: #3987e5; }
}
</style>
"""


def stat_card(icon, label, value, role="neutral"):
    return f"""
    <div class="stat-card role-{role}">
        <div class="stat-icon">{icon}</div>
        <div class="stat-label">{label}</div>
        <div class="stat-value">{value}</div>
    </div>
    """


# ── Streamlit UI ────────────────────────────────────

st.set_page_config(page_title="Excluded Contacts Extractor", page_icon="📇", layout="wide")

current_user = require_login()
render_logout_button()

st.markdown(STYLE, unsafe_allow_html=True)

st.markdown(
    """
    <div class="ece-banner">
        <div class="ece-icon">📇</div>
        <div>
            <div class="ece-title">Excluded Contacts Extractor</div>
            <div class="ece-subtitle">
                Find Contact Us entries not present in Salesforce. A contact is matched if its
                Email is found in Salesforce, or — when no email is given — if its Phone is found instead.
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

col1, col2 = st.columns(2)

with col1:
    st.subheader("📂 Contact Files")
    contact_files = st.file_uploader(
        "Upload Contact Us file(s)",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=True,
        key="contact_files",
    )

with col2:
    st.subheader("🗂️ Salesforce File")
    sf_file = st.file_uploader(
        "Upload Salesforce file",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=False,
        key="sf_file",
    )

with st.expander("⚙️ Options"):
    include_matched = st.checkbox(
        "✅ Include a 'Matched Contacts' sheet in the download",
        value=False,
        help="The original tool only exported Excluded Contacts. "
        "Turn this on to also get a sheet of the rows that matched Salesforce.",
    )

run_clicked = st.button("▶ Run Extraction", type="primary")

if run_clicked:
    logs = []

    def log(msg):
        logs.append(msg)

    if not contact_files:
        st.warning("⚠️ Please upload at least one Contact file.")
    elif not sf_file:
        st.warning("⚠️ Please upload a Salesforce file.")
    else:
        try:
            with st.spinner("Running extraction..."):
                result = run_extraction(contact_files, sf_file, log)
        except Exception as e:
            st.error(f"❌ {e}")
            with st.expander("🧾 Log"):
                st.text("\n".join(logs))
        else:
            st.success("✅ Extraction complete!")

            total = result["contact_total"]
            matched_pct = round((result["matched_count"] / total) * 100, 2) if total else 0

            st.markdown(
                '<div class="stat-grid">'
                + stat_card("📋", "Total Contact Rows", total, "neutral")
                + stat_card("✅", "Matched", result["matched_count"], "good")
                + stat_card("⚠️", "Excluded", result["excluded_count"], "serious")
                + stat_card("🎯", "Match Rate", f"{matched_pct}%", "accent")
                + "</div>",
                unsafe_allow_html=True,
            )

            sheets = {"Excluded Contacts": result["excluded"]}
            if include_matched:
                sheets["Matched Contacts"] = result["matched"]

            raw_buffer = io.BytesIO()
            with pd.ExcelWriter(raw_buffer, engine="openpyxl") as writer:
                for sheet_name, df in sheets.items():
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

            styled_buffer = style_workbook(raw_buffer)

            st.download_button(
                label="💾 Download Excel Report",
                data=styled_buffer,
                file_name="excluded_contacts.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            tab_labels = ["📋 Excluded", "✅ Matched", "🧾 Log"] if include_matched else ["📋 Excluded", "🧾 Log"]
            tabs = st.tabs(tab_labels)

            with tabs[0]:
                st.dataframe(result["excluded"].head(200), use_container_width=True)

            if include_matched:
                with tabs[1]:
                    st.dataframe(result["matched"].head(200), use_container_width=True)
                log_tab = tabs[2]
            else:
                log_tab = tabs[1]

            with log_tab:
                st.text("\n".join(logs))
