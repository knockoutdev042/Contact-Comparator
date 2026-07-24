"""
excluded_contacts_gui.py
-------------------------
GUI tool to extract Contact Us entries NOT found in Salesforce data.
Comparison is done using BOTH:
    - Phone
    - Email

A contact is considered MATCHED if:
    Phone exists in SF
    OR
    Email exists in SF

Requirements:
    pip install pandas openpyxl xlrd

Run:
    python excluded_contacts_gui.py
"""

import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# Normalization Helpers
# ──────────────────────────────────────────────────────────────────────────────

def normalize_phone(p):

    if pd.isna(p):
        return None

    p = str(p)

    p = re.sub(r'\s+', '', p)

    p = p.replace("'", "")

    try:
        if 'E+' in p.upper():
            p = str(int(float(p)))
    except:
        pass

    digits = re.sub(r'\D', '', p)

    return f"+{digits}" if digits else None


def normalize_email(email):

    if pd.isna(email):
        return None

    email = str(email)

    # Remove spaces/newlines/tabs
    email = re.sub(r'\s+', '', email)

    # Lowercase + trim
    email = email.lower().strip()

    # Basic validation
    return email if '@' in email else None


def find_column(df, possible_names):
    for col in df.columns:
        clean_col = col.strip().lower()

        for name in possible_names:
            if clean_col == name.strip().lower():
                return col

    return None


# ──────────────────────────────────────────────────────────────────────────────
# Core Logic
# ──────────────────────────────────────────────────────────────────────────────

def run_extraction(contact_paths, sf_path, output_path, log, done_callback):
    try:
        log("Loading Contact Us files...")

        contact_dfs = []

        for path in contact_paths:
            log(f"  → {path}")

            try:
                if path.lower().endswith('.csv'):
                    try:
                        df = pd.read_csv(path, encoding='utf-8-sig')
                    except UnicodeDecodeError:
                        df = pd.read_csv(path, encoding='latin1')
                else:
                    df = pd.read_excel(path)

            except Exception as e:
                log(f"❌ Failed to read file: {e}")
                continue

            if df.empty:
                log("⚠ File is empty")
                continue

            df.columns = df.columns.str.strip()

            # Flexible column detection
            phone_col = find_column(
                df,
                ['Phone', 'Mobile', 'Phone Number', 'Contact Number']
            )

            email_col = find_column(
                df,
                ['Email', 'Email Address', 'E-mail']
            )

            if not phone_col and not email_col:
                log("❌ Skipping file: No Phone or Email column found")
                continue

            df['_phone_source_col'] = phone_col
            df['_email_source_col'] = email_col

            contact_dfs.append(df)

        if not contact_dfs:
            raise Exception("No valid Contact files found.")

        # Merge all contact files
        contact = pd.concat(contact_dfs, ignore_index=True)

        log(f"✔ Total Contact rows merged: {len(contact):,}")

        # ──────────────────────────────────────────────────────────────────────
        # Load Salesforce File
        # ──────────────────────────────────────────────────────────────────────

        log("Loading Salesforce file...")

        try:
            if sf_path.lower().endswith('.csv'):
                try:
                    sf = pd.read_csv(sf_path, encoding='utf-8-sig')
                except UnicodeDecodeError:
                    sf = pd.read_csv(sf_path, encoding='latin1')
            else:
                sf = pd.read_excel(sf_path)

        except Exception as e:
            raise Exception(f"Failed to read Salesforce file: {e}")

        sf.columns = sf.columns.str.strip()

        sf_phone_col = find_column(
            sf,
            ['Phone', 'Mobile', 'Phone Number', 'Contact Number']
        )

        sf_email_col = find_column(
            sf,
            ['Email', 'Email Address', 'E-mail']
        )

        if not sf_phone_col and not sf_email_col:
            raise Exception(
                "No Phone or Email column found in Salesforce file"
            )

        log(f"✔ Salesforce rows loaded: {len(sf):,}")

        # ──────────────────────────────────────────────────────────────────────
        # Normalize Contact Data
        # ──────────────────────────────────────────────────────────────────────

        log("Normalizing Contact data...")

        contact['_phone_norm'] = None
        contact['_email_norm'] = None

        for idx, row in contact.iterrows():

            phone_col = row.get('_phone_source_col')
            email_col = row.get('_email_source_col')

            if phone_col and phone_col in contact.columns:
                contact.at[idx, '_phone_norm'] = normalize_phone(
                    row[phone_col]
                )

            if email_col and email_col in contact.columns:
                contact.at[idx, '_email_norm'] = normalize_email(
                    row[email_col]
                )

        # ──────────────────────────────────────────────────────────────────────
        # Normalize Salesforce Data
        # ──────────────────────────────────────────────────────────────────────

        log("Normalizing Salesforce data...")

        if sf_phone_col:
            sf['_phone_norm'] = sf[sf_phone_col].apply(normalize_phone)
        else:
            sf['_phone_norm'] = None

        if sf_email_col:
            sf['_email_norm'] = sf[sf_email_col].apply(normalize_email)
        else:
            sf['_email_norm'] = None

        # Remove invalid rows
        contact = contact[
            contact['_phone_norm'].notna() |
            contact['_email_norm'].notna()
        ]

        sf = sf[
            sf['_phone_norm'].notna() |
            sf['_email_norm'].notna()
        ]

        # Remove duplicates
        contact = contact.drop_duplicates(
            subset=['_phone_norm', '_email_norm']
        )

        sf = sf.drop_duplicates(
            subset=['_phone_norm', '_email_norm']
        )

        # Create lookup sets
        sf_phones = set(sf['_phone_norm'].dropna())
        sf_emails = set(sf['_email_norm'].dropna())

        log(f"✔ Unique SF phones : {len(sf_phones):,}")
        log(f"✔ Unique SF emails : {len(sf_emails):,}")

        # ──────────────────────────────────────────────────────────────────────
        # Matching Logic
        # ──────────────────────────────────────────────────────────────────────

        log("Comparing Contact data against Salesforce...")

        # Email-first matching logic

        email_exists = contact['_email_norm'].notna()

        email_match = contact['_email_norm'].isin(sf_emails)

        phone_match = contact['_phone_norm'].isin(sf_phones)

        # If email exists -> ONLY email comparison
        # Else -> fallback to phone

        matched_mask = (
            (email_exists & email_match)
            |
            (~email_exists & phone_match)
        )

        excluded = contact[~matched_mask].copy()

        matched_count = matched_mask.sum()

        # Remove helper columns
        excluded.drop(
            columns=[
                '_phone_norm',
                '_email_norm',
                '_phone_source_col',
                '_email_source_col'
            ],
            inplace=True,
            errors='ignore'
        )

        # ──────────────────────────────────────────────────────────────────────
        # Logging
        # ──────────────────────────────────────────────────────────────────────

        log("\nResults:")
        log(f"  Total Contact rows : {len(contact):,}")
        log(f"  Matched rows       : {matched_count:,}")
        log(f"  Excluded rows      : {len(excluded):,}")

        log("\nSample Excluded Phones:")
        log(str(excluded.head(10)))

        # ──────────────────────────────────────────────────────────────────────
        # Save Output
        # ──────────────────────────────────────────────────────────────────────

        log("\nSaving output file...")

        excluded.to_excel(
            output_path,
            index=False,
            sheet_name="Excluded Contacts"
        )

        # ──────────────────────────────────────────────────────────────────────
        # Excel Formatting
        # ──────────────────────────────────────────────────────────────────────

        wb = load_workbook(output_path)
        ws = wb.active

        # Header formatting
        for cell in ws[1]:
            cell.fill = PatternFill(
                "solid",
                fgColor="C00000"
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
                name="Calibri",
                size=10
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        ws.row_dimensions[1].height = 30

        # Body formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(
                    name="Calibri",
                    size=9
                )

        ws.freeze_panes = "A2"

        # Auto-width columns
        for col in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value else 0)
                for c in col
            )

            ws.column_dimensions[
                get_column_letter(col[0].column)
            ].width = min(max_len + 2, 40)

        wb.save(output_path)

        log(f"\n✅ DONE!")
        log(f"Saved to:\n{output_path}")

        done_callback(success=True)

    except Exception as e:
        log(f"\n❌ ERROR: {e}")
        done_callback(success=False)


# ──────────────────────────────────────────────────────────────────────────────
# GUI
# ──────────────────────────────────────────────────────────────────────────────

class App(tk.Tk):

    def __init__(self):
        super().__init__()

        self.title("Excluded Contacts Extractor")
        self.resizable(True, True)

        self.configure(bg="#1a1a2e")

        self.state('zoomed')

        self.contact_path = tk.StringVar()
        self.sf_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self._build_ui()

    def _build_ui(self):

        title_frame = tk.Frame(
            self,
            bg="#16213e",
            pady=18
        )

        title_frame.pack(fill="x")

        tk.Label(
            title_frame,
            text="EXCLUDED CONTACTS EXTRACTOR",
            bg="#16213e",
            fg="#e94560",
            font=("Georgia", 15, "bold")
        ).pack()

        tk.Label(
            title_frame,
            text="Compare Contact files against Salesforce",
            bg="#16213e",
            fg="#a0a0c0",
            font=("Georgia", 9)
        ).pack()

        form = tk.Frame(
            self,
            bg="#1a1a2e",
            padx=30,
            pady=20
        )

        form.pack(fill="x")

        self._file_row(
            form,
            "Contact Files",
            self.contact_path,
            [("CSV / Excel", "*.csv *.xlsx")],
            "open"
        )

        self._file_row(
            form,
            "Salesforce File",
            self.sf_path,
            [("CSV / Excel", "*.csv *.xlsx")],
            "open"
        )

        self._file_row(
            form,
            "Save Output As",
            self.output_path,
            [("Excel files", "*.xlsx")],
            "save"
        )

        btn_frame = tk.Frame(
            self,
            bg="#1a1a2e",
            pady=10
        )

        btn_frame.pack()

        self.run_btn = tk.Button(
            btn_frame,
            text="▶ RUN EXTRACTION",
            bg="#e94560",
            fg="white",
            font=("Georgia", 11, "bold"),
            relief="flat",
            bd=0,
            padx=30,
            pady=10,
            cursor="hand2",
            command=self._start
        )

        self.run_btn.pack()

        # Progress bar
        self.progress = ttk.Progressbar(
            self,
            mode="indeterminate",
            length=400
        )

        self.progress.pack(pady=10)

        # Log box
        self.log_box = tk.Text(
            self,
            height=20,
            bg="#0f3460",
            fg="white",
            font=("Courier", 9),
            state="disabled"
        )

        self.log_box.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=20
        )

    def _file_row(
        self,
        parent,
        label,
        var,
        filetypes,
        browse_type
    ):

        row = tk.Frame(parent, bg="#1a1a2e")
        row.pack(fill="x", pady=5)

        tk.Label(
            row,
            text=label,
            bg="#1a1a2e",
            fg="#a0a0c0",
            width=20,
            anchor="w"
        ).pack(side="left")

        entry = tk.Entry(
            row,
            textvariable=var,
            bg="#0f3460",
            fg="white"
        )

        entry.pack(
            side="left",
            fill="x",
            expand=True,
            padx=5
        )

        def browse():

            if browse_type == "open":

                paths = filedialog.askopenfilenames(
                    filetypes=filetypes
                )

                if paths:
                    var.set(" | ".join(paths))

            else:

                path = filedialog.asksaveasfilename(
                    filetypes=filetypes,
                    defaultextension=".xlsx"
                )

                if path:
                    var.set(path)

        tk.Button(
            row,
            text="Browse",
            command=browse,
            bg="#e94560",
            fg="white",
            relief="flat"
        ).pack(side="left")

    def _log(self, message):

        self.log_box.config(state="normal")

        self.log_box.insert("end", message + "\n")

        self.log_box.see("end")

        self.log_box.config(state="disabled")

    def _start(self):

        contact = self.contact_path.get().strip()
        sf = self.sf_path.get().strip()
        output = self.output_path.get().strip()

        if not contact:
            messagebox.showwarning(
                "Missing Input",
                "Select Contact files."
            )
            return

        if not sf:
            messagebox.showwarning(
                "Missing Input",
                "Select Salesforce file."
            )
            return

        if not output:
            messagebox.showwarning(
                "Missing Output",
                "Select output file."
            )
            return

        contact_paths = contact.split(" | ")

        self.run_btn.config(state="disabled")

        self.progress.start(10)

        def done(success):

            self.progress.stop()

            self.run_btn.config(state="normal")

        threading.Thread(
            target=run_extraction,
            args=(
                contact_paths,
                sf,
                output,
                self._log,
                done
            ),
            daemon=True
        ).start()


if __name__ == "__main__":
    app = App()
    app.mainloop()